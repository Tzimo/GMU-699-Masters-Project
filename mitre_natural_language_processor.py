# Import packages
import os, datetime, nltk, docx, pptx, PyPDF2, sys, subprocess, tkinter as tk, openpyxl, ntpath, textract

from tkinter.filedialog import askopenfilename
from openpyxl import Workbook


# Datasets
raw_sentences = []




# Functino to read .pdf, .txt, .doc, .docx files
def read_file(filename):

    if filename.endswith('.txt'):
        f = open(filename, 'rb')
        for line in f:
            print(line)
            raw_sentences.append(line.strip('\n'))
    elif filename.endswith('.docx'):
        doc = docx.Document(filename)
        for line in doc.paragraphs:
            raw_sentences.append(line.text)
    elif filename.endswith('.doc'):
        # placeholder
        print('processing .doc')   
    elif filename.endswith('.pdf'):
        # placeholder
        print('Processing .pdf')
    

# Write sentences to Excel for classification
def write_to_excel(filename, processed_data):
    wb = Workbook()
    ws = wb.active
    
    # Create sheet
    #ws1 = wb.create_sheet(ntpath.basename(filename),0)
    print(ntpath.basename(filename))
    #ws.title(str(ntpath.basename(filename)))

    # for row in ws.iter_rows(min_row=1, max_col=1, max_row=len(processed_data)):
    #    for cell in row:
    #ws.append(processed_data)
    for i in range(1, len(processed_data)):
        d = ws.cell(row=i, column=1)
        d.value = processed_data[i]
    
    # Save Excel
    wb.save(str(os.path.splitext(filename)[0] + '_processed.xlsx'))

def main():
    current_directory = os.path.dirname(os.path.realpath(__file__))

    #filename = 'C:\\Users\\Michael\\Documents\\GMU\\SYST 699 Masters Project\\GMU-699-Masters-Project\\Tzimourakas_Paper_Review1.docx'
    #filename = 'C:\\Users\\Michael\\Documents\\GMU\\SYST 699 Masters Project\\GMU-699-Masters-Project\\Tzimourakas_Paper_Review1.txt'

    if len(sys.argv) !=2:
        # Open gui to select file
        filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
    else:
        # Get file
        filename = str(sys.argv[1])
    
   # print(ntpath.basename(filename))

    # Call read_file function
    read_file(filename)

    # Filter out empty sentences
    processed_sentences = list(filter(None, raw_sentences))


    # Write data to Excel
    write_to_excel(filename, processed_sentences)
    #print(len(processed_sentences))

    #for sentence in processed_sentences:
    #    print(sentence)

if __name__ == '__main__':
    main()
